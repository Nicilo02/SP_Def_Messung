import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl import load_workbook
from copy import copy

# Pfad zur Quelldatei, aus der die Kopf- und Fußzeilen kopiert werden
source_file = "Excelfiles/Verarbeitung.xlsx"  # Quelldatei mit den Kopf- und Fußzeilen

# Öffnen der Quelldatei
source_wb = openpyxl.load_workbook(source_file)
source_ws = source_wb.active

# Erstellen einer neuen Excel-Datei
new_wb = openpyxl.Workbook()
new_ws = new_wb.active

# Kopieren der Kopf- und Fußzeilen von der Quelldatei zur neuen Datei
new_ws.oddHeader.left.text = source_ws.oddHeader.left.text
new_ws.oddHeader.center.text = source_ws.oddHeader.center.text
new_ws.evenHeader.center.text = source_ws.evenHeader.center.text

# Fußzeilen
new_ws.oddFooter.center.text = source_ws.oddFooter.center.text
new_ws.evenFooter.center.text = source_ws.evenFooter.center.text
new_ws.oddFooter.right.text = source_ws.oddFooter.right.text

new_ws.oddFooter.left.text = 'Seite &P'

# Speichern der neuen Excel-Datei mit den kopierten Kopf- und Fußzeilen
new_wb.save("Excelfiles/Neu_2.xlsx")

def copy_styles(source_file, target_file):
    # Lade Arbeitsmappen
    source_wb = load_workbook(source_file)
    target_wb = load_workbook(target_file)
    
    # Übernehmen Sie das erste Arbeitsblatt (Sie können dies anpassen)
    source_ws = source_wb.active
    target_ws = target_wb.active
    
    for row in source_ws.iter_rows():
        for cell in row:
            target_cell = target_ws[cell.coordinate]
            
            # Kopiere den Wert
            target_cell.value = cell.value
            
            # Kopiere den Stil (verwende `copy` für Stilobjekte)
            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.border = copy(cell.border)
                target_cell.fill = copy(cell.fill)
                target_cell.number_format = copy(cell.number_format)
                target_cell.protection = copy(cell.protection)
                target_cell.alignment = copy(cell.alignment)
    
    # Speichere die Änderungen
    target_wb.save('Neu.xlsx')

# Beispielaufruf
source_file = "Excelfiles/Verarbeitung.xlsx"
target_file = "Neu_2.xlsx"
copy_styles(source_file, target_file)


# Funktion ausführen
copy_styles(source_file, target_file)







